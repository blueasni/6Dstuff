#from functools import singledispatchmethod
from sys import byteorder
from numpy.core.arrayprint import set_legacy_print_mode
from numpy.core.numerictypes import english_capitalize
from pandas import DataFrame, ExcelWriter, to_datetime
from datetime import date, datetime  
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from pandas._libs.tslibs.timestamps import Timestamp
import openpyxl
from openpyxl import load_workbook
import math
from PIL import ImageFont
from openpyxl import Workbook
import pandas as pd
from openpyxl.styles import Alignment,Border,Color,PatternFill,Side, alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from  openpyxl.worksheet.dimensions import ColumnDimension,DimensionHolder
#import seaborn as sns
def Send_Dataframe_To_Excel(dataframe):
        columns_of_type_datetime = list(dataframe.select_dtypes(include=["datetime", "datetime64", "datetime64[ns]", "datetimetz"]).columns)
            #find all columns of type datetime
            
        for column in columns_of_type_datetime :
            dataframe[column] = to_datetime(dataframe[column], format='%Y-%m-%d %H:%M:%S', errors='coerce').dt.date
                #convert them to pandas datetime
                
        with ExcelWriter("output_file.xlsx", engine="openpyxl", date_format='DD/MM/YYYY', datetime_format='DD/MM/YYYY') as writer:
            dataframe.to_excel(writer, sheet_name="Sheet1", index=False)

dataframe = DataFrame({'ID': [1, 2, 3, 4, 5, 6, 7],
            'text': ['text1', 'text2', 'text3', 'text4', 'text5', 'text6', 'text7'],
            'number': [11, 12, 13, 14, 15, 16, 17,],
            'date': [Timestamp('2011-01-01 00:20:00'), Timestamp('2012-02-02 00:00:00'), Timestamp('2013-03-03 00:00:00'), Timestamp('2014-04-04 00:00:00'), Timestamp('2015-05-05 00:00:00'), Timestamp('2016-06-06 00:00:00'), Timestamp('2017-07-07 00:00:00')]})
#csv_df = pd.read_csv('employee.csv')
#csv_df.style.set_caption("test employee.csv")           
#Send_Dataframe_To_Excel(dataframe)

df1 = pd.DataFrame([["AAA", "BBB"]], columns=["Spam", "Egg"])  
df2 = pd.DataFrame([["ABC", "XYZ"]], columns=["Foo", "Bar"])  
df = pd.DataFrame(
    [
        [date(2014, 1, 31), date(1999, 9, 24)],
        [datetime(1998, 5, 26, 23, 33, 4), datetime(2014, 2, 28, 13, 5, 13)],
    ],
    index=["Date", "Datetime"],
    columns=["X", "Y"],
)  
with pd.ExcelWriter("path_to_file.xlsx", date_format="YYYY-MM-DD", datetime_format="YYYY-MM-DD") as writer:
    df.to_excel(writer) 
#with pd.ExcelWriter("multi_sheet.xlsx",if_sheet_exists="replace") as writer1:
#    df1.to_excel(writer1, sheet_name="Sheet1")  
#    df2.to_excel(writer1, sheet_name="Sheet2")

with ExcelWriter("df_in_single_sheet.xlsx",engine="openpyxl",if_sheet_exists="overlay") as writer:
    df1.to_excel(writer, sheet_name="Sheet1", index=False)
    #ws = writer.book.get_worksheet_by_name('Sheet1')
    #ws = writer._get_sheet_name("Sheet1")
    #ws.write('A1', 'This is a caption')
    df2.to_excel(writer, sheet_name="Sheet1", index=False, startcol=5) 
    df2.to_excel(writer, sheet_name="Sheet1", index=False, startcol=10) 
#--------------------------------------------------------
wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(csv_df, index=False, header=True):
    ws.append(r)
style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
table = Table(displayName="penguins", ref="A1:" + get_column_letter(ws.max_column) + str(ws.max_row))
print(get_column_letter(ws.max_column))
print(str(ws.max_row))
rng = ws["A1:"+ get_column_letter(ws.max_column) + str(ws.max_row)]
thin=Side(border_style="thin",color="000000")
singleBorder= Side(border_style="thin",color="ff0000")
alignment = Alignment(horizontal="center",vertical="center",indent=1)
border = Border(top=singleBorder,left=singleBorder,right=singleBorder,bottom=singleBorder)
for row in rng:
    for c in row:
        c.border = border
        c.alignment=alignment
        
rng = ws['A1':'C2']
#=============
'''
column_widths = []
for row in ws.iter_rows():
    for i, cell in enumerate(row):
        try:
            column_widths[i] = max(column_widths[i], len(str(cell.value)))
        except IndexError:
            column_widths.append(len(str(cell.value)))

for i, column_width in enumerate(column_widths):
    ws.column_dimensions[get_column_letter(i + 1)].width = column_width+2
'''
def auto_width(workb,ws): #adjusting auto column width
    column_widths = []
    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
            except IndexError:
                column_widths.append(len(str(cell.value)))
    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width + 2
    workb.save("table-pandas.xlsx")
#++++++++++++++
def adjust_excel_column_widths(worksheet, fontname):
    
    #font = ImageFont.truetype("calibri",11,encoding="unicode") 
    font = ImageFont.truetype('home/blue/Downloads/nerd font/AnonymousPro/AnonymiceProNerdFont-Bold.ttf',int(11 * 4/3),encoding="unicode")
    excel_unit_width = font.getlength('0')

    def getwidth(cell_value):
        return 0.0 if cell_value is None else font.getlength(str(cell_value))

    column_number = 0
    for column in worksheet.columns:
        column_number += 1
        column_letter = get_column_letter(column_number)
        column_width = math.ceil( max(getwidth(cell.value) for cell in column) / excel_unit_width )
        worksheet.column_dimensöions[column_letter].width = column_width
#---------------------------------------------------------------------
table.tableStyleInfo = style
ws.add_table(table)
auto_width(wb,ws)
#adjust_excel_column_widths(ws, '/home/blue/openpxl/export-dataframes-to-excel/calibri.ttf')
wb.save("table-pandas.xlsx")
#===================================================

