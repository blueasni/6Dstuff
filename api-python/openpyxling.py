# Read the existing workbook with openpyxl
# path = "C:\\Users\\admin\\Documents\\MEDIUM\\Python-to-Excel'\\Example_Report.xlsx" # EDIT THE PATH TO AN EXCEL WORKBOOK IN YOUR OWN SCRIPT
path = r"C:\Users\admin\Documents\MEDIUM\Python-to-Excel\Example_Report.xlsx"

workbook = load_workbook(path, 
                         read_only=False, 
                         keep_vba=False) # set to True if reading a macro-enabled workbook

# Create a Pandas Excel writer object using openpyxl as the engine
writer = pd.ExcelWriter(path, 
                        engine='openpyxl')
writer.book = workbook

# Store the existing sheets to a dictionary of {title: worksheet object} pairs
writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)

# Select a worksheet to edit
ws = workbook['Report']

# Clear any pre-existing values within the following cell ranges
for row in ws['B3:E1000']:
    for cell in row:
        cell.value = None
        
# Convert the dataframe to openpyxl Excel object within the workbook
df.to_excel(writer, 
            sheet_name='Report', 
            float_format='%0.2f', 
            header=False, 
            index=False, 
            startrow=2, 
            startcol=1, 
            engine='openpyxl')

# Font formatting
font = Font(name='Arial',
            size=11,
            bold=True,
            color='404161') #HEX code

# Cell background Color
background_fill = PatternFill(start_color='F5E5B7', #HEX code
                               fill_type='solid')

# Border formatting
border = Border(left=Side(style='medium', color=colors.BLACK),
                right=Side(style='medium', color=colors.BLACK),
                top=Side(style='medium', color=colors.BLACK),
                bottom=Side(style='medium', color=colors.BLACK))

# Alignment formatting
alignment = Alignment(horizontal='center',
                      vertical='top',
                      wrapText=True)

# Define cell type (i.e. "General", "Text", etc) 
number_format = numbers.FORMAT_DATE_XLSX14 #mm-dd-yy

# Define the cell range to edit
cell_range = ws['B3:B17']

# Apply the formatting to the cell range
for row in cell_range:
    for cell in row:
        cell.font = font
        cell.fill = background_fill
        cell.border = border
        cell.alignment = alignment
        cell.number_format = number_format
        
chart_img = Image(img='chart_output.png')
chart_img.anchor = 'G4'
ws.add_image(chart_img)

# Save the workbook and close the writer
writer.save()
writer.close()
#---------------------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Border, Side

wb = Workbook()
ws = wb.active
rng = ws['B2':'D3']
s = Side(style = 'double', color = '0000ff')
for row in rng:
    for c in row:
        c.value = 123
        c.border = Border(left = s, right = s, top = s, bottom = s) 
wb.save('border.xlsx')
#---------------------------------------------------------------------------------
   def is_merged_horizontally(cell):
        """
        Checks if cell is merged horizontally with an another cell
        @param cell: cell to check
        @return: True if cell is merged horizontally with an another cell, else False
        """
        cell_coor = cell.coordinate
        if cell_coor not in ws.merged_cells:
            return False
        for rng in ws.merged_cells.ranges:
            if cell_coor in rng and len(list(rng.cols)) > 1:
                return True
        return False

    for col_number, col in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(col_number)

        max_length = max(
            len(str(cell.value or "")) for cell in col if not is_merged_horizontally(cell)
        )
        adjusted_width = (max_length + 2) * 0.95
        ws.column_dimensions[col_letter].width = adjusted_width
#===================================================================================
