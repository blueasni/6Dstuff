from openpyxl.cell.cell import WriteOnlyCell
import pandas as pd
from openpyxl import load_workbook,Workbook,cell
from openpyxl.utils.dataframe import dataframe_to_rows
wb = Workbook(write_only=True)
ws = wb.create_sheet()

cell = WriteOnlyCell(ws)
cell.style = 'Pandas'
df = pd.read_excel("data.xlsx") 
def format_first_row(row, cell):
   for c in row:
       cell.value = c
       yield cell

rows = dataframe_to_rows(df)
first_row = format_first_row(next(rows), cell)
ws.append(first_row)

for row in rows:
    row = list(row)
    cell.value = row[0]
    row[0] = cell
    ws.append(row)

wb.save("openpyxl_stream.xlsx")
wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)

for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'

wb.save("pandas_openpyxl-1.xlsx")