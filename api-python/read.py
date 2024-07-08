#from functools import singledispatchmethod
import os
from sys import byteorder
from openpyxl.compat import numbers
from sqlalchemy import URL, create_engine
from pandas import DataFrame, ExcelWriter, to_datetime
from datetime import date, datetime  
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from pandas._libs.tslibs.timestamps import Timestamp
import openpyxl
from openpyxl import load_workbook,Workbook,cell
import math
from PIL import ImageFont
import pandas as pd
from openpyxl.styles import Alignment,Border,Color, Font,PatternFill,Side, alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension,DimensionHolder
from sqlalchemy import text

#api_data = pd.read_csv('ChangeSim.txt')

#order_id = api_data['order_id']
#print(order_id)
#order_id = (str(list(order_id.values))).strip('[]')
#print(str(order_id))
order = pd.read_csv('csv.csv')

for row in order.columns:
    #print(len(order[row]))
    for i in row.it:
        print(order[row].iloc[i])
    #print()

