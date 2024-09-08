from sys import byteorder
from openpyxl.compat import numbers
from sqlalchemy import URL, create_engine
from pandas import DataFrame, ExcelWriter, to_datetime
from datetime import date, timedelta  
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from pandas._libs.tslibs.timestamps import Timestamp 
from openpyxl import load_workbook,Workbook,cell
import math,datetime,openpyxl,os,pandas as pd
from PIL import ImageFont
from openpyxl.styles import Alignment,Border,Color, Font,PatternFill,Side, alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension,DimensionHolder
from sqlalchemy import text
from pygnuplot import gnuplot
thin=Side(border_style="thin",color="000000")
test = Side(border_style="double",color="ffffff")
testBorder = Border(top=test,bottom=test,left=test,right=test)
singleBorder= Side(border_style="thin",color="121212")
doubleBorder = Side(border_style="double",color="008000")
alignment = Alignment(horizontal="center",vertical="center",indent=1)
border = Border(top=singleBorder,left=singleBorder,right=singleBorder,bottom=singleBorder)

url_object = URL.create("mysql+pymysql",username="rptuserc1_ro",password="bssuser@6Dtech",host="10.3.74.22",database="WKN_COM",port=6033,)
url_object_2 = URL.create("mysql+pymysql",username="rptuserc2_ro",password="bssuser@6Dtech",host="10.3.74.22",database="WKN_COM",port=6044,)
url_object_3 = URL.create("mysql+pymysql",username="rptuserc3_ro",password="bssuser@6Dtech",host="10.3.74.22",database="WKN_COM",port=6055,)

engine = create_engine(url_object)
connection = engine.connect()
PARTITION = 'p'+str(date.today().month)
TODAY = date.today()
YESTERDAY = date.today() - timedelta(days=1)
#q=text("SELECT A.order_id, B.SUB_ORDER_ID,B.service_id, A.order_type, B.sub_order_state, B.state_reason, A.created_date,	A.CHANNEL, 'Asnake' As Name_,'2024-06-24 06:30:00' AS St_TIme, CASE WHEN B.state_reason LIKE '400 :: Service Details already exist for serviceId : %' THEN 'The Service is already active' WHEN B.state_reason = '1503 :: Service limit reached for a profile :: Billing' THEN 'limit reached' WHEN B.state_reason = '404 :: Blank Sim Details is not available in db :: NMS' THEN 'Duplicate request' WHEN B.state_reason = '404 :: SIM Details is not valid for pairing :: NMS' THEN 'Duplicate request' WHEN B.state_reason = '400 :: SIM is already paired with another msisdn :: NMS' THEN 'IMSI already paired' WHEN B.state_reason = '400 :: MSISDN is already paired with another SIM :: NMS' THEN 'IMSI already paired' WHEN B.state_reason = '404 :: Sim Details is not available in DB :: NMS' THEN 'ICCID already paired' WHEN B.state_reason = '424 :: Read timed out while invoking third party :: ERP' THEN 'Raised to Tibco' WHEN B.state_reason LIKE 'Object uid=%' THEN 'Duplicate Request' WHEN B.state_reason LIKE '555 :: To borrow%Birr you need to have spent at least%Birr within the last 30 days :: NCC' THEN 'Not eligible' WHEN B.state_reason = 'Please attach the document' THEN 'The document is rejected' WHEN B.state_reason = '400 :: Invalid Patch for Path:/identities for operation:replace. The Identity supplied in Where clause is not associated with this device. :: NCC' THEN 'Duplicate request' WHEN B.state_reason LIKE '658 :: This operation is not allowed since customer has an active loan ::%' THEN 'Customer have active loan' WHEN B.state_reason = '400 :: DNA Pairing Failed! :: NMS' THEN 'Duplicate order' ELSE null END AS DESCRIPTION, CASE WHEN DESCRIPTION IS NOT NULL THEN 'NO' END AS WA FROM COM_ORDER_MASTER PARTITION(p6) A, COM_SUB_ORDER_DETAILS PARTITION(p6) B WHERE B.SUB_ORDER_STATE != 'Completed' AND B.state_reason != 'Waiting for Bank Callback' AND B.Sub_order_state != 'In-Progress' AND A.ORDER_TYPE in ('Onboarding', 'AddService', 'TransferOfService', 'ChangeSim', 'TerminateService', 'ChangeSubscription', 'UpdateStarterPackKYC', 'HardUnbarring', 'HardBarring', 'SoftUnbarring', 'SoftBarring', 'AddServiceToNewAccount', 'ConnectionMigration') AND A.CREATED_DATE BETWEEN '2024-06-23 21:30:00' and '2024-06-24 06:30:00' AND A.ORDER_ID = B.ORDER_ID ORDER BY A.CREATED_DATE;")
#q=text("SELECT DATE(ORDER_DATE),ORDER_STATE,COUNT(*) FROM COM_ORDER_MASTER WHERE DATE(ORDER_DATE)>='2024-07-01' and ORDER_TYPE='onboarding' and ORDER_STATE in ('Completed','Failed')  GROUP BY DATE(ORDER_DATE),ORDER_STATE;")
q=text("SELECT DATE(ORDER_DATE) AS ORDER_DATE,ORDER_STATE,coalesce(COUNT(ORDER_STATE), 0) FROM COM_ORDER_MASTER WHERE CREATED_DATE BETWEEN '2024-07-03' and '2024-07-04' and ORDER_STATE in ('Completed','Failed') GROUP BY DATE(ORDER_DATE),ORDER_STATE")
TOP5_FAILED_ORDERS = text("SELECT DATE(A.CREATED_DATE) as Created_Date ,NAME,count(*) from COM_ORDER_MASTER PARTITION({0}) A,COM_ORDER_STAGES PARTITION({1}) B where A.CREATED_DATE BETWEEN {2} and {3} and B.STATE='Failed' and A.ORDER_ID=B.ORDER_ID group by DATE(A.CREATED_DATE),NAME  order by DATE(A.CREATED_DATE),count(*) desc;".format(PARTITION,PARTITION,YESTERDAY,TODAY))
ONBOARDING = text("SELECT DATE(ORDER_DATE) AS ORDER_DATE,ORDER_STATE,COUNT(*) AS COUNT FROM COM_ORDER_MASTER WHERE DATE(ORDER_DATE)>='2024-07-02' and ORDER_TYPE='onboarding' and ORDER_STATE in ('Completed','Failed')  GROUP BY DATE(ORDER_DATE),ORDER_STATE;")
COMPLETED = text("SELECT (SELECT count(*) FROM patients WHERE order_state='Completed' AND DATE(ORDER_DATE)>='2024-07-02' AND ORDER_TYPE='onboarding') AS Completed, (SELECT count(*) FROM patients WHERE order_state='Failed' AND DATE(ORDER_DATE)>='2024-07-02' AND ORDER_TYPE='onboarding') AS Failed")
ALL_ORDERS = text("SELECT DATE(ORDER_DATE) AS ORDERED_DATE, ORDER_STATE,ORDER_TYPE, COUNT(*) AS COUNT FROM COM_ORDER_MASTER WHERE ORDER_TYPE IN ('Onboarding','Addservice','AddServiceToNewAccount','AddSubscription','BlockVoucher','BookDeposit','AdjustMainAccount','CancelSubscription','ChangeSim','ChangeSubscription','CreateDocument','CreateIdentification','Gifting','HardBarring','LifeCycleSync','LifeCycleSyncTermination','LineBarring','LineUnBarring','MakePayment','MoveToFWA','NumberRecycle','ResumeService''StopAutoRenewal','SuspendService','TransferOfService','UpdateBucket','UpdateCreditLimit','UpdateLanguage','UpdateProfile','UnlockMpesa','UpdateService','DeviceBlacklistWhitelist','VoucherRecharge') AND ORDER_STATE IN ('Failed', 'Completed') AND ORDER_DATE >= {0} AND ORDER_DATE < {1} GROUP BY ORDERED_DATE, ORDER_STATE,ORDER_TYPE".format(YESTERDAY,TODAY))
ALL_ORDERS_TEST = text("SELECT DATE(ORDER_DATE) AS ORDERED_DATE, ORDER_STATE,ORDER_TYPE, COUNT(*) AS COUNT FROM COM_ORDER_MASTER WHERE ORDER_TYPE IN ('Onboarding','Addservice') AND ORDER_STATE IN ('Failed', 'Completed') AND ORDER_DATE BETWEEN {0} and {1} GROUP BY ORDERED_DATE, ORDER_STATE,ORDER_TYPE;".format(YESTERDAY,TODAY))
ORDER_TYPE = ['AddService','AddSubscription','CancelSubscription','ChangeSim','Onboarding','StopAutoRenewal','UpdateLanguage','AddServiceToNewAccount','AdjustMainAccount','BlockVoucher','BookDeposit','ChangeSubscription','CreateDocument','CreateIdentification','DeviceBlacklistWhitelist','Gifting','HardBarring','LifeCycleSync','LifeCycleSyncTermination','LineBarring','LineUnBarring','LockMpesa','MakePayment','MoveToFWA','NumberRecycle','ResumeService','SuspendService','TransferOfService','UpdateBucket','UpdateProfile','UnlockMpesa','UpdateCreditLimit','UpdateService','VoucherRecharge']

TEST = text("SELECT DATE(ORDER_DATE) AS ORDERED_DATE, ORDER_TYPE, ORDER_STATE, COUNT(*) AS COUNT FROM COM_ORDER_MASTER WHERE ORDER_TYPE IN ('AddService','AddSubscription','CancelSubscription','ChangeSim','Onboarding','StopAutoRenewal','UpdateLanguage','AddServiceToNewAccount','AdjustMainAccount','BlockVoucher','BookDeposit','ChangeSubscription','CreateDocument','CreateIdentification','DeviceBlacklistWhitelist','Gifting','HardBarring','LifeCycleSync','LifeCycleSyncTermination','LineBarring','LineUnBarring','LockMpesa','MakePayment','MoveToFWA','NumberRecycle','ResumeService','SuspendService','TransferOfService','UpdateBucket','UpdateProfile','UnlockMpesa','UpdateCreditLimit','UpdateService','VoucherRecharge') AND ORDER_STATE IN ('Failed', 'Completed') AND ORDER_DATE >= '%s' AND ORDER_DATE < '%s' GROUP BY ORDERED_DATE, ORDER_STATE,ORDER_TYPE;" % (YESTERDAY,TODAY))
ALL_API_QUERY=text("select APINAME,count(*) from APIGW_RPT where DATE(PROCESS_DATE)=DATE(CURRENT_DATE)-1 group by APINAME order by APINAME;")
TOP5_ALL_API_QUERY=text("select APINAME,count(*) from APIGW_RPT where DATE(PROCESS_DATE)=date(current_date)-1 and APINAME in ('QueryPlan','getBasicDetails','searchNumber','ProductOffering','add-subscription') group by APINAME;")
order_columns = ['Order_Date','Order_State','Onboarding','AddService','AddServiceToNewAccount','AddSubscription','BlockVoucher','BookDeposit','AdjustMainAccount','CancelSubscription','ChangeSim','ChangeSubscription','CreateDocument','CreateIdentification','Gifting','HardBarring','LifeCycleSync','LifeCycleSyncTermination','LineBarring','LineUnBarring','MakePayment','MoveToFWA','NumberRecycle','ResumeService''StopAutoRenewal','SuspendService','TransferOfService','UpdateBucket','UpdateCreditLimit','UpdateLanguage','UpdateProfile','UnlockMpesa','UpdateService','DeviceBlacklistWhitelist','VoucherRecharge']
df = pd.DataFrame(columns=order_columns)
FINAL_ORDER = pd.DataFrame()
copier = pd.DataFrame()
api_data = pd.read_sql(TEST,connection) 
#ALL_API = pd.read_sql(ALL_API_QUERY,connection) 
#TOP5_ALL_API = pd.read_sql(TOP5_ALL_API_QUERY,connection) 
api_data_opt = pd.DataFrame
unique = pd.DataFrame
TRAV = []
VAL = []
#api_data = pd.read_excel("data.xlsx") 
#api_data = pd.read_csv("data.csv")
count_row = api_data.shape[0]  # Gives number of rows
count_col = api_data.shape[1]  # Gives number of columns
(r,c) = api_data.shape # r,c row and column length of api_data df
#result = api_data[api_data['ORDER_TYPE'] == "AddService"]
#print(result.iloc[0,0]," ",result.iloc[0,1]," ",result.iloc[0,2]," ",result.iloc[0,3])
def auto_width(workb,sheet): #adjusting auto column width
    column_widths = []
    for row_1 in sheet.iter_rows():
        for k, cell2 in enumerate(row_1):
            try:
                column_widths[k] = max(column_widths[k], len(str(cell2.value)))
            except IndexError:
                column_widths.append(len(str(cell2.value)))
    for n, column_width in enumerate(column_widths):
        sheet.column_dimensions[get_column_letter(n + 1)].width = column_width + 2
    #workb.save("topOrder.xlsx")
    #os.startfile("topOrder.xlsx")
def do_border(sheet):
    area = sheet["A1:"+ get_column_letter(sheet.max_column) + str(sheet.max_row)]
    for row in area:
        for cell in row:
            cell.border = border
            cell.alignment=alignment
def sheet_append(sheet, data, hd):
    for r in dataframe_to_rows(data, index=False, header=hd):
        sheet.append(r)
def sheet_noheader(sheet,data):
    COM_FAIL = []
    col_a = sheet.column_dimensions['A']
    col_a.fill = PatternFill(fill_type='solid', fgColor='C0C0C0')
    count = 1
    mod_order_type=ORDER_TYPE
    for x in range(2, len(ORDER_TYPE)*2, 2):
        sheet.merge_cells(start_row=1, start_column=x, end_row=1, end_column=x+1)
        sheet[get_column_letter(x)+str(1)] = mod_order_type[count]
        count +=1
    mod_order_type.insert(0," ")
    for r in range(len(ORDER_TYPE)*2):
        if (r % 2) == 0:
            COM_FAIL.append('C')
        else:
            COM_FAIL.append('F')
    COM_FAIL.insert(0," ")
    sheet.append(COM_FAIL)
    for r in dataframe_to_rows(data, index=False, header=False):
        sheet.append(r)
def do_orderBreakup():
    global TRAV, VAL
    TRAV = TOTAL_BUP.drop(TOTAL_BUP.index[2],axis=0)
    TRAV = TRAV.drop(TRAV.columns[[1]],axis = 1)
    VAL = []
    for i in range(TRAV.shape[1]):
        VAL.append(TRAV.iloc[0,i])
        VAL.append(TRAV.iloc[1,i])
    VAL.insert(0,YESTERDAY)
def assemble_finalResult():
    cols = list(api_data.columns)
    for row in range(count_row):
        for i in range(count_col):
            #order_date = api_data.iloc[row,0]
            #order_type = api_data.iloc[row,1]
            #order_state = api_data.iloc[row,2]
            #state_count = api_data.iloc[row,3]
            row_num = FINAL_ORDER[FINAL_ORDER['ORDER_TYPE'] == api_data.iloc[row,1]].index[0]
            FINAL_ORDER['ORDER_DATE'] = YESTERDAY
            #FINAL_ORDER['ORDER_DATE'] = pd.Series([YESTERDAY for x in range(len(FINAL_ORDER.index))])
            FINAL_ORDER.at[row_num, api_data.iloc[row,2]] = api_data.iloc[row,3]
            FINAL_ORDER.at[row_num, 'TOTAL'] = int(FINAL_ORDER.iloc[row_num,2]) + int(FINAL_ORDER.iloc[row_num,3])

def copier():
    global copier,api_data_opt
    orderdate = ['%s' % YESTERDAY for s in ORDER_TYPE]
    compl, failed, compl = [0] * len(ORDER_TYPE),[0] * len(ORDER_TYPE),[0] * len(ORDER_TYPE)
    col_header = {'ORDER_DATE' : orderdate, 'ORDER_TYPE' : ORDER_TYPE, 'Completed' : compl, 'Failed' : failed}
    copier = pd.DataFrame(col_header)
    api_data_opt = pd.DataFrame({'ORDER_DATE' : [], 'ORDER_TYPE' : [], 'Completed' : [], 'Failed' : []})    
    for i in range(api_data.shape[0]):
        od = api_data.iloc[i , 0]
        ot = api_data.iloc[i , 1]
        os = api_data.iloc[i , 2]
        co = api_data.iloc[i , 3]       
        if os == 'Completed':
            os = co
            co = 0
        else:
            os = 0
        list = [od, ot,os,co]
        #api_data_opt.loc[len(api_data_opt)] = list
        #api_data_opt.append({list})

    duplicated = api_data_opt[api_data_opt.duplicated("ORDER_TYPE")]
    count_r = duplicated.shape[0]
    unique = api_data_opt.drop_duplicates(subset=["ORDER_TYPE"]).reset_index(drop=True)
    for r in range(count_r):
        ot = duplicated.iloc[r,1]
        oc = duplicated.iloc[r,2]
        of = duplicated.iloc[r,3]
        result = unique[unique['ORDER_TYPE'] == ot]
        print(result)
        unique.iat[unique.loc[unique['ORDER_TYPE'] == ot].index[0], unique.columns.get_loc('Completed')] = unique.iat[unique.loc[unique['ORDER_TYPE'] == ot].index[0], unique.columns.get_loc('Completed')] + oc
        unique.iat[unique.loc[unique['ORDER_TYPE'] == ot].index[0], unique.columns.get_loc('Failed')] = unique.iat[unique.loc[unique['ORDER_TYPE'] == ot].index[0], unique.columns.get_loc('Failed')] + of
def fill_zero():
    global FINAL_ORDER
    ORDER_DATE, COMPLETED, FAILED, TOTAL = [0] * len(ORDER_TYPE),[0] * len(ORDER_TYPE),[0] * len(ORDER_TYPE),[0] * len(ORDER_TYPE)
    DICT = {'ORDER_DATE' : ORDER_DATE,'ORDER_TYPE': ORDER_TYPE, 'Completed': COMPLETED, 'Failed': FAILED, 'TOTAL' : TOTAL}
    FINAL_ORDER = pd.DataFrame(DICT)
def onboarding():
    onboard = FINAL_ORDER[FINAL_ORDER['ORDER_TYPE'] == 'Onboarding']
    onb = pd.DataFrame()
    for r in range(len(onboard)):
        DICT = {'ORDER_DATE' : ['Completed','Failed'],onboard.iloc[r,0]:[onboard.iloc[r,2],onboard.iloc[r,3]]}
        onb = pd.DataFrame(DICT)
    return onb
def total_orders():
    completed_sum = 0
    failed_sum=0
    for r in range(len(FINAL_ORDER)):
        completed_sum = completed_sum + int(FINAL_ORDER.iloc[r,2])
        failed_sum = failed_sum + int(FINAL_ORDER.iloc[r,3])
    DICT = {'ORDER_DATE' : ['Completed','Failed'],FINAL_ORDER.iloc[r,0]:[completed_sum,failed_sum]}
    total = pd.DataFrame(DICT)
    return total
copier()
fill_zero()
assemble_finalResult()
#api_data[['Created_Date']] = api_data[['Created_Date']].astype(str)
#api_data[['order_id', 'SUB_ORDER_ID']] = api_data[['order_id', 'SUB_ORDER_ID']].astype(str)
#api_data = api_data.iloc[1:]
#writer = pd.ExcelWriter('pandas_multiple.xlsx', engine='xlsxwriter')
#api_data.to_excel(writer, sheet_name='Sheet1')  # Default position, cell A1.
#api_data.to_excel(writer, sheet_name='Sheet1', startcol=4)
#api_data.to_excel(writer, sheet_name='Sheet1', startcol=7)

#--------------------------------------------------------
wb = Workbook()
#ws = wb.create_sheet('All_orders')
top5failed = wb.create_sheet('TOP 5 Failed Orders')
total_breakup_sheet = wb.create_sheet('Total_Orders_Breakup')
org_order_bup = wb.create_sheet('Total Orders Breakup(org)')
test_sheet = wb.create_sheet('Test_sheet')
total_Orders_breakup = wb.create_sheet('Total Orders breakup')
transpo = wb.create_sheet('Transpose')
cop = wb.create_sheet('Copier')
onboard_sheet = wb.create_sheet('Onboarding')
fine = wb.create_sheet('Final All Orders')
tot_orders = wb.create_sheet('Total Orders')
all_api_sheet = wb.create_sheet('ALL API')
top5_api_sheet = wb.create_sheet('Top 5 API')
ws = wb.active

ws.title = "All_orders"
sheet_append(ws, FINAL_ORDER,True)
do_border(ws)
auto_width(wb,ws)
TOTAL_BUP = FINAL_ORDER.transpose()

TOP5FAILED = FINAL_ORDER.nlargest(5,'Failed').drop(['Completed', 'TOTAL'], axis=1)
TOTAL_BUP = FINAL_ORDER.set_index('ORDER_TYPE').T.drop(['ORDER_DATE'],axis=0)

do_orderBreakup()
TOTAL_BUP.insert(0, 'ORDER_DATE', [YESTERDAY,YESTERDAY,'TOTAL'])
sheet_noheader(org_order_bup, TOTAL_BUP)
auto_width(wb,org_order_bup)
do_border(org_order_bup)

sheet_append(total_breakup_sheet, TOTAL_BUP,True)
do_border(total_breakup_sheet)
auto_width(wb,total_breakup_sheet)

sheet_append(top5failed,TOP5FAILED,True)
do_border(top5failed)
auto_width(wb,top5failed)
top5failed.merge_cells(start_row=2, start_column=1, end_row=6, end_column=1)

auto_width(wb,test_sheet)
do_border(test_sheet)
sheet_append(test_sheet,TRAV,False)

auto_width(wb,transpo)
do_border(transpo)
sheet_append(transpo,api_data,True)

auto_width(wb,cop)
do_border(cop)
#sheet_append(cop,unique,True)

sheet_noheader(total_Orders_breakup,pd.DataFrame({"order" : VAL}).transpose())
auto_width(wb,total_Orders_breakup)
do_border(total_Orders_breakup)

sheet_append(fine,FINAL_ORDER,True)
auto_width(wb,fine)
do_border(fine)

#onboarding(onboard_sheet)
sheet_append(onboard_sheet,onboarding(),True)
auto_width(wb,onboard_sheet)
do_border(onboard_sheet)

sheet_append(tot_orders,total_orders(),True)
auto_width(wb,tot_orders)
do_border(tot_orders)

'''
sheet_append(all_api_sheet,total_orders(),True)
auto_width(wb,all_api_sheet)
do_border(all_api_sheet)

sheet_append(top5_api_sheet,total_orders(),True)
auto_width(wb,top5_api_sheet)
do_border(top5_api_sheet)
'''

for row in range(2 , ws.max_row):
    ell = ws['A'+str(row)]
    ell.font = Font(name='Calibri',size=11,bold=True,italic=False, color='008000')
    ell.number_format = "@"
for i in range(1 , ws.max_column+1):
    ell = ws[get_column_letter(i)+str(1)]
    ell.font = Font(name='Calibri',size=11,bold=True,italic=False, color='008000')
    ell.fill = PatternFill(fill_type='solid',start_color='FFFF00',end_color='FFFF00')
    ell.border = Border(top=thin , bottom = thin , left = thin , right = thin)
    ell.number_format = "@"
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=True)
table = Table(displayName="API_TEST", ref="A1:" + get_column_letter(ws.max_column) + str(ws.max_row))

rng = ws["A1:"+ get_column_letter(ws.max_column) + str(ws.max_row)]
for row in rng:
    for c in row:
        c.border = border
        c.alignment=alignment
#ws.column_dimensions.group('A', 'D', hidden=True,outline_level=1)        
ws.column_dimensions.group('F', 'F', hidden=True)        
ws.cell(2,2).number_format='@'

#++++++++++++++
col_range = ws['A3:A4']
for row in col_range:
    for c in row:
        c.fill = PatternFill(fill_type='solid',start_color='FFFF00',end_color='FFFF00')
table.tableStyleInfo = style
ws.add_table(table)
#auto_width(wb,ws)
#adjust_excel_column_widths(ws, '/home/blue/openpxl/export-dataframes-to-excel/calibri.ttf')
c = ws['A2']
ws.freeze_panes = c
wb.save("api-test.xlsx")
os.startfile("api-test.xlsx")
#connection.close
